from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, status, Form, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import select
from openpyxl import load_workbook
from typing import List, Dict

from app.core.deps import role_required
from app.db.session import get_db
from app.db.models_catalogs import Cliente, Proveedor, TipoRecurso, Recurso, Especialidad, Unidad
from app.schemas.catalogs import (
    ClienteCreate,
    ClienteRead,
    ProveedorCreate,
    ProveedorRead,
    TipoRecursoCreate,
    TipoRecursoRead,
    RecursoCreate,
    RecursoUpdate,
    RecursoRead,
    EspecialidadCreate,
    EspecialidadRead,
    UnidadCreate,
    UnidadRead,
)
from app.services.excel_recursos import generar_plantilla_excel, procesar_excel_recursos


router = APIRouter(prefix="/catalogos", tags=["catalogos"]) 


# Clientes
@router.get("/clientes", response_model=list[ClienteRead])
def list_clientes(db: Session = Depends(get_db), _: None = Depends(role_required(["Administrador", "Cotizador"]))):
    return list(db.scalars(select(Cliente)).all())


@router.post("/clientes", response_model=ClienteRead, status_code=status.HTTP_201_CREATED)
def create_cliente(payload: ClienteCreate, db: Session = Depends(get_db), _: None = Depends(role_required(["Administrador"]))):
    c = Cliente(**payload.model_dump())
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


# Proveedores
@router.get("/proveedores", response_model=list[ProveedorRead])
def list_proveedores(db: Session = Depends(get_db), _: None = Depends(role_required(["Administrador", "Cotizador"]))):
    return list(db.scalars(select(Proveedor)).all())


@router.post("/proveedores", response_model=ProveedorRead, status_code=status.HTTP_201_CREATED)
def create_proveedor(payload: ProveedorCreate, db: Session = Depends(get_db), _: None = Depends(role_required(["Administrador"]))):
    p = Proveedor(**payload.model_dump())
    db.add(p)
    db.commit()
    db.refresh(p)
    return p


# Tipos de Recurso
@router.get("/tipos_recurso", response_model=list[TipoRecursoRead])
def list_tipos(db: Session = Depends(get_db)):
    return list(db.scalars(select(TipoRecurso)).all())


@router.post("/tipos_recurso", response_model=TipoRecursoRead, status_code=status.HTTP_201_CREATED)
def create_tipo(nombre: str = Form(...), db: Session = Depends(get_db)):
    existing = db.scalar(select(TipoRecurso).where(TipoRecurso.nombre == nombre))
    if existing:
        raise HTTPException(status_code=400, detail="El tipo ya existe")
    t = TipoRecurso(nombre=nombre)
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


# Recursos
@router.get("/recursos")
def list_recursos(db: Session = Depends(get_db)):
    recursos = db.scalars(select(Recurso)).all()
    result = []
    for recurso in recursos:
        unidad_obj = db.get(Unidad, recurso.id_unidad) if recurso.id_unidad else None
        result.append({
            "id_recurso": recurso.id_recurso,
            "id_tipo_recurso": recurso.id_tipo_recurso,
            "descripcion": recurso.descripcion,
            "id_unidad": recurso.id_unidad,
            "cantidad": float(recurso.cantidad),
            "costo_unitario_predeterminado": float(recurso.costo_unitario_predeterminado),
            "costo_total": float(recurso.costo_total),
            "id_proveedor_preferido": recurso.id_proveedor_preferido,
            "atributos": recurso.atributos,
            "unidad": unidad_obj.nombre if unidad_obj else ""
        })
    return result


@router.post("/recursos", status_code=status.HTTP_201_CREATED)
def create_recurso(payload: RecursoCreate, db: Session = Depends(get_db)):
    # Crear el recurso en la base de datos
    r = Recurso(**payload.model_dump())
    db.add(r)
    db.commit()
    db.refresh(r)
    
    # Obtener el nombre de la unidad para devolverlo
    unidad_obj = db.get(Unidad, r.id_unidad) if r.id_unidad else None
    
    return {
        "id_recurso": r.id_recurso,
        "id_tipo_recurso": r.id_tipo_recurso,
        "descripcion": r.descripcion,
        "id_unidad": r.id_unidad,
        "cantidad": float(r.cantidad),
        "costo_unitario_predeterminado": float(r.costo_unitario_predeterminado),
        "costo_total": float(r.costo_total),
        "id_proveedor_preferido": r.id_proveedor_preferido,
        "atributos": r.atributos,
        "unidad": unidad_obj.nombre if unidad_obj else ""
    }


@router.patch("/recursos/{id_recurso}")
def update_recurso(id_recurso: int, payload: RecursoUpdate, db: Session = Depends(get_db)):
    r = db.get(Recurso, id_recurso)
    if not r:
        raise HTTPException(status_code=404, detail="Recurso no encontrado")
    data = payload.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(r, k, v)
    db.commit()
    db.refresh(r)
    
    # Obtener el nombre de la unidad para devolverlo
    unidad_obj = db.get(Unidad, r.id_unidad) if r.id_unidad else None
    
    return {
        "id_recurso": r.id_recurso,
        "id_tipo_recurso": r.id_tipo_recurso,
        "descripcion": r.descripcion,
        "id_unidad": r.id_unidad,
        "cantidad": float(r.cantidad),
        "costo_unitario_predeterminado": float(r.costo_unitario_predeterminado),
        "costo_total": float(r.costo_total),
        "id_proveedor_preferido": r.id_proveedor_preferido,
        "atributos": r.atributos,
        "unidad": unidad_obj.nombre if unidad_obj else ""
    }


# Carga masiva desde Excel (hoja activa con columnas: descripcion, tipo, unidad, costo + columnas adicionales)
@router.post("/recursos/carga_masiva")
def carga_masiva(file: UploadFile = File(...), db: Session = Depends(get_db), _: None = Depends(role_required(["Administrador"]))):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Archivo Excel inválido")
    wb = load_workbook(file.file, read_only=True, data_only=True)
    ws = wb.active
    
    # Obtener todos los headers de la primera fila
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    expected_basic = ["descripcion", "tipo", "unidad", "costo"]
    
    # Verificar que las primeras 4 columnas sean las esperadas
    if headers[:4] != expected_basic:
        raise HTTPException(status_code=400, detail=f"Las primeras 4 columnas deben ser: {expected_basic}")
    
    # Obtener columnas adicionales (después de las 4 básicas)
    additional_columns = headers[4:] if len(headers) > 4 else []

    tipos_cache: dict[str, int] = {}
    count_inserted = 0
    for row in ws.iter_rows(min_row=2):
        descripcion = (row[0].value or "").strip()
        tipo = (row[1].value or "").strip()
        unidad = (row[2].value or "").strip()
        costo = float(row[3].value or 0)
        
        if not descripcion or not tipo or not unidad:
            continue
            
        # Crear diccionario de atributos adicionales
        atributos = {}
        for i, col_name in enumerate(additional_columns):
            if col_name and i + 4 < len(row):
                cell_value = row[i + 4].value
                if cell_value is not None:
                    # Convertir a string y limpiar
                    atributos[col_name] = str(cell_value).strip()
        
        if tipo not in tipos_cache:
            existing = db.scalar(select(TipoRecurso).where(TipoRecurso.nombre == tipo))
            if not existing:
                existing = TipoRecurso(nombre=tipo)
                db.add(existing)
                db.commit()
                db.refresh(existing)
            tipos_cache[tipo] = existing.id_tipo_recurso
            
        recurso = Recurso(
            id_tipo_recurso=tipos_cache[tipo],
            descripcion=descripcion,
            unidad=unidad,
            costo_unitario_predeterminado=costo,
            atributos=atributos if atributos else None
        )
        db.add(recurso)
        count_inserted += 1
        if count_inserted % 200 == 0:
            db.commit()

    db.commit()
    return {"insertados": count_inserted}


# Especialidades
@router.get("/especialidades", response_model=list[EspecialidadRead])
def listar_especialidades(db: Session = Depends(get_db)):
    """Listar todas las especialidades"""
    return list(db.scalars(select(Especialidad).order_by(Especialidad.nombre)).all())


@router.post("/especialidades", response_model=EspecialidadRead, status_code=status.HTTP_201_CREATED)
def crear_especialidad(
    nombre: str = Form(...), 
    descripcion: str = Form(""), 
    db: Session = Depends(get_db)
):
    """Crear una nueva especialidad"""
    # Verificar si ya existe
    existing = db.scalar(select(Especialidad).where(Especialidad.nombre == nombre))
    if existing:
        raise HTTPException(status_code=400, detail="La especialidad ya existe")
    
    especialidad = Especialidad(nombre=nombre, descripcion=descripcion if descripcion else None)
    db.add(especialidad)
    db.commit()
    db.refresh(especialidad)
    return especialidad


# Unidades
@router.get("/unidades", response_model=list[UnidadRead])
def listar_unidades(db: Session = Depends(get_db)):
    """Listar todas las unidades"""
    return list(db.scalars(select(Unidad).order_by(Unidad.nombre)).all())


@router.post("/unidades", response_model=UnidadRead, status_code=status.HTTP_201_CREATED)
def crear_unidad(
    nombre: str = Form(...), 
    simbolo: str = Form(""), 
    descripcion: str = Form(""), 
    db: Session = Depends(get_db)
):
    """Crear una nueva unidad"""
    # Verificar si ya existe
    existing = db.scalar(select(Unidad).where(Unidad.nombre == nombre))
    if existing:
        raise HTTPException(status_code=400, detail="La unidad ya existe")
    
    unidad = Unidad(
        nombre=nombre,
        simbolo=simbolo if simbolo else None,
        descripcion=descripcion if descripcion else None
    )
    db.add(unidad)
    db.commit()
    db.refresh(unidad)
    return unidad


# Excel - Generar plantilla para carga de recursos
@router.post("/recursos/generar-plantilla-excel")
def generar_plantilla(
    atributos: List[Dict[str, str]] = Body(...),
    nombre_planilla: str = Body(...),
):
    """
    Genera un archivo Excel con una tabla formateada para carga de recursos.
    
    Body esperado:
    {
        "atributos": [
            {"nombre": "Descripción", "tipo": "texto"},
            {"nombre": "Unidad", "tipo": "texto"},
            {"nombre": "Cantidad", "tipo": "entero"},
            {"nombre": "Costo Unitario", "tipo": "numerico"}
        ],
        "nombre_planilla": "Personal"
    }
    """
    try:
        excel_file = generar_plantilla_excel(atributos, nombre_planilla)
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=Plantilla_{nombre_planilla.replace(' ', '_')}.xlsx"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando plantilla: {str(e)}")


# Excel - Cargar recursos desde Excel
@router.post("/recursos/cargar-desde-excel")
async def cargar_recursos_excel(
    file: UploadFile = File(...),
    id_tipo_recurso: int = Form(...),
    atributos: str = Form(...),  # JSON string
    db: Session = Depends(get_db)
):
    """
    Procesa un archivo Excel y carga recursos en la base de datos.
    
    El archivo debe tener:
    - Fila 3: Encabezados (nombres de atributos)
    - Fila 4+: Datos de recursos
    """
    import json
    
    try:
        # Parsear atributos
        atributos_list = json.loads(atributos)
        print(f"DEBUG RUTA: Atributos recibidos: {atributos_list}")
        print(f"DEBUG RUTA: ID tipo recurso: {id_tipo_recurso}")
        
        # Leer contenido del archivo
        file_content = await file.read()
        print(f"DEBUG RUTA: Archivo leído, tamaño: {len(file_content)} bytes")
        
        # Procesar Excel
        resultado = procesar_excel_recursos(file_content, id_tipo_recurso, atributos_list)
        print(f"DEBUG RUTA: Resultado procesamiento: {resultado['total_procesados']} recursos, {len(resultado['errores'])} errores")
        
        if resultado['errores'] and len(resultado['recursos']) == 0:
            return {
                'success': False,
                'errores': resultado['errores'],
                'total_procesados': 0
            }
        
        # Guardar recursos en BD
        recursos_guardados = 0
        recursos_actualizados = 0
        errores_guardado = []
        ids_recursos_procesados = []  # Para devolver los IDs
        
        print(f"DEBUG RUTA: Iniciando guardado de {len(resultado['recursos'])} recursos")
        
        for idx, recurso_data in enumerate(resultado['recursos'], 1):
            print(f"DEBUG RUTA: Procesando recurso {idx}: {recurso_data.get('descripcion')}")
            
            # Buscar unidad por nombre
            unidad_nombre = recurso_data.get('unidad', '')
            unidad = db.scalar(select(Unidad).where(Unidad.nombre == unidad_nombre))
            
            if not unidad:
                # Si la unidad no existe, crearla automáticamente
                print(f"DEBUG RUTA: Unidad '{unidad_nombre}' no encontrada, creándola automáticamente...")
                nueva_unidad = Unidad(
                    nombre=unidad_nombre,
                    simbolo=unidad_nombre,
                    descripcion=f"Unidad {unidad_nombre} (creada automáticamente)"
                )
                db.add(nueva_unidad)
                db.flush()  # Para obtener el ID
                unidad = nueva_unidad
                print(f"DEBUG RUTA: Unidad '{unidad_nombre}' creada con ID: {unidad.id_unidad}")
            else:
                print(f"DEBUG RUTA: Unidad encontrada: {unidad.nombre} (ID: {unidad.id_unidad})")
            
            # Extraer atributos personalizados
            atributos_personalizados = {}
            campos_base = ['descripcion', 'unidad', 'cantidad', 'costo_unitario', 'costo_total', 'id_tipo_recurso']
            
            for key, value in recurso_data.items():
                if key not in campos_base:
                    atributos_personalizados[key] = value
            
            # Verificar si ya existe un recurso con la misma descripción Y los mismos atributos
            # Buscar todos los recursos con la misma descripción
            recursos_con_misma_desc = db.scalars(
                select(Recurso).where(
                    Recurso.id_tipo_recurso == id_tipo_recurso,
                    Recurso.descripcion == recurso_data.get('descripcion')
                )
            ).all()
            
            # Buscar uno que tenga exactamente los mismos atributos personalizados
            recurso_existente = None
            for r in recursos_con_misma_desc:
                # Comparar atributos personalizados
                atributos_existentes = r.atributos or {}
                
                # Si ambos tienen los mismos atributos (mismo contenido), es el mismo recurso
                if atributos_existentes == atributos_personalizados:
                    recurso_existente = r
                    break
            
            if recurso_existente:
                # Actualizar
                print(f"DEBUG RUTA: Actualizando recurso existente ID: {recurso_existente.id_recurso}")
                recurso_existente.id_unidad = unidad.id_unidad
                recurso_existente.cantidad = recurso_data.get('cantidad', 0)
                recurso_existente.costo_unitario_predeterminado = recurso_data.get('costo_unitario', 0)
                recurso_existente.costo_total = recurso_data.get('costo_total', 0)
                recurso_existente.atributos = atributos_personalizados if atributos_personalizados else None
                recursos_actualizados += 1
                ids_recursos_procesados.append(recurso_existente.id_recurso)
            else:
                # Crear nuevo
                print(f"DEBUG RUTA: Creando nuevo recurso: {recurso_data.get('descripcion')}")
                nuevo_recurso = Recurso(
                    id_tipo_recurso=id_tipo_recurso,
                    descripcion=recurso_data.get('descripcion'),
                    id_unidad=unidad.id_unidad,
                    cantidad=recurso_data.get('cantidad', 0),
                    costo_unitario_predeterminado=recurso_data.get('costo_unitario', 0),
                    costo_total=recurso_data.get('costo_total', 0),
                    atributos=atributos_personalizados if atributos_personalizados else None
                )
                db.add(nuevo_recurso)
                db.flush()  # Para obtener el ID antes de commit
                recursos_guardados += 1
                ids_recursos_procesados.append(nuevo_recurso.id_recurso)
                print(f"DEBUG RUTA: Nuevo recurso creado con ID: {nuevo_recurso.id_recurso}")
        
        print(f"DEBUG RUTA: Haciendo commit...")
        db.commit()
        print(f"DEBUG RUTA: Commit exitoso. Total guardados: {recursos_guardados}, actualizados: {recursos_actualizados}")
        
        todos_errores = resultado['errores'] + errores_guardado
        
        # Si hay errores, mostrarlos pero aún así retornar success si se guardó al menos uno
        if todos_errores and (recursos_guardados + recursos_actualizados) == 0:
            return {
                'success': False,
                'recursos_guardados': 0,
                'recursos_actualizados': 0,
                'total_procesados': 0,
                'ids_recursos_procesados': [],
                'errores': todos_errores
            }
        
        return {
            'success': True,
            'recursos_guardados': recursos_guardados,
            'recursos_actualizados': recursos_actualizados,
            'total_procesados': recursos_guardados + recursos_actualizados,
            'ids_recursos_procesados': ids_recursos_procesados,
            'errores': todos_errores
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")



"""
Servicio para generar y procesar archivos Excel de recursos
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict, Any


def generar_plantilla_excel(atributos: List[Dict[str, str]], nombre_planilla: str) -> BytesIO:
    """
    Genera un archivo Excel con una tabla formateada para carga de recursos.
    
    Args:
        atributos: Lista de diccionarios con {'nombre': str, 'tipo': str}
        nombre_planilla: Nombre de la planilla de recursos
    
    Returns:
        BytesIO con el contenido del archivo Excel
    """
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Recursos"
    
    # Configurar anchos de columna
    for col_idx, attr in enumerate(atributos, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Side(border_style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # Título de la planilla
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(atributos))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"PLANILLA DE RECURSOS - {nombre_planilla.upper()}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Instrucciones
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(atributos))
    instruction_cell = ws.cell(row=2, column=1)
    instruction_cell.value = "Complete los datos de cada recurso en las filas siguientes. No modifique los encabezados."
    instruction_cell.font = Font(italic=True, size=10, color="666666")
    instruction_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25
    
    # Encabezados de columnas (fila 3)
    for col_idx, attr in enumerate(atributos, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = attr['nombre']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    ws.row_dimensions[3].height = 25
    
    # Agregar filas de ejemplo con validación de tipo
    example_data = {
        'descripcion': 'SUPERVISOR DE OBRA',
        'unidad': 'HH',
        'cantidad': 40,
        'costo_unitario': 21.50,
        'costo_total': 860.00,
    }
    
    # Fila de ejemplo (fila 4)
    for col_idx, attr in enumerate(atributos, start=1):
        cell = ws.cell(row=4, column=col_idx)
        attr_id = attr['nombre'].lower().replace(' ', '_')
        
        # Mapear nombres a IDs
        if 'descripci' in attr_id or 'detalle' in attr_id:
            cell.value = example_data.get('descripcion', 'Ejemplo')
        elif 'unidad' in attr_id:
            cell.value = example_data.get('unidad', 'un')
        elif 'cantidad' in attr_id:
            cell.value = example_data.get('cantidad', 1)
        elif 'unitario' in attr_id or 'precio' in attr_id:
            cell.value = example_data.get('costo_unitario', 0.0)
        elif 'total' in attr_id:
            cell.value = example_data.get('costo_total', 0.0)
        else:
            cell.value = f"Ejemplo {attr['nombre']}"
        
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        
        # Formatear números
        if attr['tipo'] in ['numerico', 'entero']:
            if 'costo' in attr_id or 'precio' in attr_id or 'total' in attr_id:
                cell.number_format = '$#,##0.00'
            elif attr['tipo'] == 'entero':
                cell.number_format = '0'
            else:
                cell.number_format = '#,##0.00'
    
    # Agregar 20 filas vacías para que el usuario complete
    for row_idx in range(5, 25):
        for col_idx, attr in enumerate(atributos, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Valores por defecto
            if attr['tipo'] in ['numerico', 'entero']:
                if 'costo' in attr['nombre'].lower() or 'precio' in attr['nombre'].lower():
                    cell.number_format = '$#,##0.00'
                elif attr['tipo'] == 'entero':
                    cell.number_format = '0'
                else:
                    cell.number_format = '#,##0.00'
    
    # Congelar paneles (primera fila de encabezados)
    ws.freeze_panes = 'A4'
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def procesar_excel_recursos(
    file_content: bytes, 
    id_tipo_recurso: int,
    atributos_esperados: List[Dict[str, str]]
) -> Dict[str, Any]:
    """
    Procesa un archivo Excel de recursos y retorna los datos para guardar en BD.
    
    Args:
        file_content: Contenido del archivo Excel
        id_tipo_recurso: ID del tipo de recurso (planilla)
        atributos_esperados: Lista de atributos con nombre y tipo
    
    Returns:
        Dict con 'recursos' (lista) y 'errores' (lista)
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(BytesIO(file_content), data_only=True)
    ws = wb.active
    
    # Buscar encabezados en diferentes filas (1, 2, 3)
    headers = []
    header_row = 0
    
    for row_candidate in [1, 2, 3]:
        temp_headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_candidate, column=col_idx).value
            if cell_value and str(cell_value).strip():
                temp_headers.append(str(cell_value).strip())
            else:
                break
        
        # Si encontramos encabezados válidos, usarlos
        if len(temp_headers) >= 2:  # Mínimo 2 columnas
            headers = temp_headers
            header_row = row_candidate
            break
    
    if len(headers) == 0:
        return {'recursos': [], 'errores': ['No se encontraron encabezados válidos en el archivo']}
    
    print(f"DEBUG SERVICIO: Encabezados encontrados en fila {header_row}: {headers}")
    
    # Función auxiliar para normalizar texto (quitar tildes, espacios, minúsculas)
    def normalizar_texto(texto):
        import unicodedata
        # Quitar tildes
        texto_sin_tildes = ''.join(
            c for c in unicodedata.normalize('NFD', texto)
            if unicodedata.category(c) != 'Mn'
        )
        # Minúsculas y reemplazar espacios por _
        return texto_sin_tildes.lower().replace(' ', '_')
    
    # Mapear nombres de columnas a IDs de atributos
    # Si un header no está en atributos_esperados, lo tratamos como atributo personalizado (texto)
    columna_a_atributo = {}
    for idx, header in enumerate(headers):
        # Normalizar el header
        header_normalizado = normalizar_texto(header)
        
        # Buscar coincidencia con atributos esperados
        encontrado = False
        for attr in atributos_esperados:
            attr_normalizado = normalizar_texto(attr['nombre'])
            
            # Comparar ambos normalizados
            if attr_normalizado == header_normalizado:
                columna_a_atributo[idx] = attr
                encontrado = True
                break
        
        # Si no se encontró, asumimos que es un atributo personalizado de tipo texto
        if not encontrado:
            columna_a_atributo[idx] = {
                'nombre': header_normalizado,
                'tipo': 'texto'
            }
    
    # Procesar filas de datos (desde la fila siguiente a los encabezados)
    start_data_row = header_row + 1
    recursos = []
    errores = []
    
    print(f"DEBUG SERVICIO: Iniciando procesamiento de filas desde fila {start_data_row} hasta {ws.max_row}")
    print(f"DEBUG SERVICIO: Headers encontrados: {headers}")
    print(f"DEBUG SERVICIO: Mapeo de columnas: {columna_a_atributo}")
    
    for row_idx in range(start_data_row, ws.max_row + 1):
        # Leer datos de la fila
        row_data = {}
        fila_vacia = True
        
        for col_idx in range(len(headers)):
            cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
            
            if cell_value is not None and str(cell_value).strip():
                fila_vacia = False
                
                if col_idx in columna_a_atributo:
                    attr = columna_a_atributo[col_idx]
                    # Usar la función normalizar_texto para el attr_id también
                    attr_id = normalizar_texto(attr['nombre'])
                    
                    # Convertir según tipo
                    try:
                        if attr['tipo'] == 'entero':
                            row_data[attr_id] = int(float(cell_value))
                        elif attr['tipo'] == 'numerico':
                            row_data[attr_id] = float(cell_value)
                        else:
                            row_data[attr_id] = str(cell_value).strip()
                    except (ValueError, TypeError):
                        errores.append(f"Fila {row_idx}, columna '{attr['nombre']}': valor inválido '{cell_value}'")
                        continue
        
        # Si la fila no está vacía, agregar recurso
        if not fila_vacia and row_data:
            print(f"DEBUG SERVICIO: Fila {row_idx} - Datos: {row_data}")
            
            # Validar campos requeridos
            campos_requeridos = ['descripcion', 'unidad', 'cantidad', 'costo_unitario']
            faltantes = [c for c in campos_requeridos if c not in row_data or not row_data[c]]
            
            if faltantes:
                print(f"DEBUG SERVICIO: Fila {row_idx}: faltan campos requeridos {faltantes}")
                errores.append(f"Fila {row_idx}: faltan campos requeridos {faltantes}")
                continue
            
            # Calcular costo total si no existe
            if 'costo_total' not in row_data:
                cantidad = row_data.get('cantidad', 0)
                costo_unitario = row_data.get('costo_unitario', 0)
                row_data['costo_total'] = cantidad * costo_unitario
            
            row_data['id_tipo_recurso'] = id_tipo_recurso
            recursos.append(row_data)
            print(f"DEBUG SERVICIO: Recurso agregado: {row_data.get('descripcion')}")
    
    print(f"DEBUG SERVICIO: Total recursos procesados: {len(recursos)}, errores: {len(errores)}")
    
    return {
        'recursos': recursos,
        'errores': errores,
        'total_procesados': len(recursos)
    }


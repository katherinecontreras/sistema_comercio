from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Literal, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.db.models import Material, TipoMaterial, ensure_total_cantidad_struct


HeaderKind = Literal["base", "atribute"]

# Número máximo de filas con fórmulas (hasta fila 1000)
MAX_FORMULA_ROWS = 1000

BASE_TITLE_FIELD_MAP: Dict[str, str] = {
    "detalle": "detalle",
    "cantidad": "cantidad",
    "unidad": "unidad",
    "$unitario": "costo_unitario",
    "$ total": "costo_total",
    "$total": "costo_total",
}

BASE_ID_FIELD_FALLBACK: Dict[int, str] = {
    1: "detalle",
    2: "cantidad",
    3: "unidad",
    4: "costo_unitario",
    5: "costo_total",
}

NUMERIC_BASE_IDS = {2, 4, 5}
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
TITLE_FILL_COLOR = "e0f2fe"
HEADER_FILL_COLOR = "e0f2fe"
TOTAL_LABEL_FILL_COLOR = "e0f2fe"
TOTAL_VALUE_FILL_COLOR = "f8fafc"


@dataclass
class HeaderSpec:
    kind: HeaderKind
    header_id: int
    titulo: str
    calculo: Dict[str, Any]
    is_cantidad: bool
    order: int
    raw: Dict[str, Any]

    @property
    def numeric_hint(self) -> bool:
        if self.kind == "base":
            return self.header_id in NUMERIC_BASE_IDS
        return bool(self.is_cantidad or (self.calculo or {}).get("activo"))


def _normalize_header_type(raw_type: Optional[str]) -> HeaderKind:
    if not raw_type:
        return "base"
    lowered = raw_type.lower()
    return "atribute" if lowered.startswith("atr") else "base"


def _resolve_base_field(header: Dict[str, Any]) -> Optional[str]:
    titulo = (header.get("titulo") or "").strip().lower()
    if titulo in BASE_TITLE_FIELD_MAP:
        return BASE_TITLE_FIELD_MAP[titulo]
    base_id = header.get("id_header_base")
    if isinstance(base_id, int):
        return BASE_ID_FIELD_FALLBACK.get(base_id)
    return None


def _build_ordered_headers(tipo: TipoMaterial) -> List[HeaderSpec]:
    headers: List[HeaderSpec] = []
    seen: set[Tuple[HeaderKind, int]] = set()

    base_headers = list(tipo.headers_base or [])
    attr_headers = list(tipo.headers_atributes or [])

    base_map = {hb["id_header_base"]: hb for hb in base_headers if "id_header_base" in hb}
    attr_map = {
        ha["id_header_atribute"]: ha for ha in attr_headers if "id_header_atribute" in ha
    }

    order_entries = sorted(tipo.order_headers or [], key=lambda entry: entry.get("order", 0))

    def add_header(kind: HeaderKind, header_dict: Dict[str, Any]) -> None:
        header_id = int(
            header_dict["id_header_base"]
            if kind == "base"
            else header_dict["id_header_atribute"]
        )
        if (kind, header_id) in seen:
            return
        if kind == "base" and not (header_dict.get("active", True)):
            return

        titulo = (header_dict.get("titulo") or "").strip() or (
            header_dict.get("titulo_default") or ""
        )
        calculo = header_dict.get("calculo") or {}
        is_cantidad = bool(header_dict.get("isCantidad", False))
        raw_order = header_dict.get("order")
        fallback_order = (
            999 if (kind == "base" and header_id == 5) else header_id
        )
        spec = HeaderSpec(
            kind=kind,
            header_id=header_id,
            titulo=titulo,
            calculo=calculo,
            is_cantidad=is_cantidad,
            order=int(raw_order if raw_order is not None else fallback_order),
            raw=header_dict,
        )
        headers.append(spec)
        seen.add((kind, header_id))

    for entry in order_entries:
        header_id = entry.get("id")
        if header_id is None:
            continue
        kind = _normalize_header_type(entry.get("type"))
        if kind == "base":
            header_dict = base_map.get(int(header_id))
            if header_dict:
                add_header("base", header_dict)
        else:
            header_dict = attr_map.get(int(header_id))
            if header_dict:
                add_header("atribute", header_dict)

    remaining_base = [
        hb
        for hb in base_headers
        if (("base", hb.get("id_header_base")) not in seen)
        and hb.get("active", True)
    ]
    remaining_base.sort(
        key=lambda hb: (hb.get("order") or (999 if hb.get("id_header_base") == 5 else hb.get("id_header_base", 999)))
    )
    for hb in remaining_base:
        add_header("base", hb)

    remaining_attr = [
        ha
        for ha in attr_headers
        if ("atribute", ha.get("id_header_atribute")) not in seen
    ]
    remaining_attr.sort(
        key=lambda ha: ha.get("order") or ha.get("id_header_atribute", 999)
    )
    for ha in remaining_attr:
        add_header("atribute", ha)

    headers.sort(key=lambda spec: spec.order)
    return headers


def _safe_to_float(value: Any) -> Optional[float]:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return 0.0
        normalized = stripped.replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return 0.0
    return 0.0


def _gather_operation_values(
    resolve_value: Any,
    operacion: Dict[str, Any],
) -> List[float]:
    valores: List[float] = []
    for header_id in operacion.get("headers_base") or []:
        resolved = resolve_value("base", int(header_id))
        numeric = _safe_to_float(resolved)
        valores.append(numeric)
    for header_id in operacion.get("headers_atributes") or []:
        resolved = resolve_value("atribute", int(header_id))
        numeric = _safe_to_float(resolved)
        valores.append(numeric)
    return valores


def _compute_operation_result(operator: str, valores: Iterable[float]) -> Optional[float]:
    valores_list = list(valores)
    if not valores_list:
        return None
    operator = operator.lower()
    if operator == "multiplicacion":
        result = 1.0
        for value in valores_list:
            result *= value
        return result
    if operator == "division":
        result = valores_list[0]
        for divisor in valores_list[1:]:
            if divisor == 0:
                return None
            result /= divisor
        return result
    if operator == "suma":
        return sum(valores_list)
    if operator == "resta":
        result = valores_list[0]
        for value in valores_list[1:]:
            result -= value
        return result
    return None


def _calculate_calculo(
    headers_lookup: Dict[Tuple[HeaderKind, int], HeaderSpec],
    resolve_value: Any,
    header: HeaderSpec,
) -> Optional[float]:
    calculo = header.calculo or {}
    if not calculo.get("activo"):
        return None
    operaciones = calculo.get("operaciones") or []
    if not operaciones:
        return None

    ops_iterable = operaciones if calculo.get("isMultiple") else operaciones[:1]
    resultado: Optional[float] = None

    for operacion in ops_iterable:
        operator = (operacion.get("tipo") or "multiplicacion").lower()
        valores = _gather_operation_values(resolve_value, operacion)
        operation_value = _compute_operation_result(operator, valores)
        if operation_value is None:
            continue
        if resultado is None:
            resultado = operation_value
            continue
        if operator == "multiplicacion":
            resultado *= operation_value
        elif operator == "division":
            if operation_value == 0:
                return None
            resultado /= operation_value
        elif operator == "suma":
            resultado += operation_value
        elif operator == "resta":
            resultado -= operation_value

    return resultado


def _get_raw_base_value(material: Material, header_dict: Dict[str, Any]) -> Any:
    field_name = _resolve_base_field(header_dict)
    if not field_name:
        return None
    return getattr(material, field_name, None)


def _build_row_values(tipo: TipoMaterial, material: Material, headers: List[HeaderSpec]) -> Dict[Tuple[HeaderKind, int], Any]:
    headers_lookup = {(spec.kind, spec.header_id): spec for spec in headers}

    attr_values = {
        int(attr["id_header_atribute"]): attr.get("value")
        for attr in (material.atributos or [])
        if "id_header_atribute" in attr
    }

    memo: Dict[Tuple[HeaderKind, int], Any] = {}
    visiting: set[Tuple[HeaderKind, int]] = set()

    base_lookup = {
        spec.header_id: spec.raw for spec in headers if spec.kind == "base"
    }
    attr_lookup = {
        spec.header_id: spec.raw for spec in headers if spec.kind == "atribute"
    }

    def resolve(kind: HeaderKind, header_id: int) -> Any:
        key = (kind, header_id)
        if key in memo:
            return memo[key]
        if key in visiting:
            return None
        visiting.add(key)

        if kind == "base":
            header_dict = base_lookup.get(header_id)
            spec = headers_lookup.get(key)
            raw_value = _get_raw_base_value(material, header_dict or {})
            value: Any = raw_value
            if spec and spec.calculo:
                calculo_value = _calculate_calculo(headers_lookup, resolve, spec)
                if calculo_value is not None:
                    value = calculo_value
        else:
            spec = headers_lookup.get(key)
            raw_value = attr_values.get(header_id)
            value = raw_value
            if spec and spec.calculo:
                calculo_value = _calculate_calculo(headers_lookup, resolve, spec)
                if calculo_value is not None:
                    value = calculo_value

        visiting.remove(key)
        memo[key] = value
        return value

    for spec in headers:
        resolve(spec.kind, spec.header_id)

    return memo


def _apply_title_style(cell) -> None:
    cell.font = Font(color="000000", bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(fill_type="solid", fgColor=TITLE_FILL_COLOR)
    cell.border = THIN_BORDER


def _apply_header_style(cell) -> None:
    cell.font = Font(color="000000", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL_COLOR)
    cell.border = THIN_BORDER


def _apply_body_style(cell) -> None:
    cell.font = Font(color="000000")
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER


def _auto_adjust_columns(worksheet, column_count: int) -> None:
    for col_idx in range(1, column_count + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in worksheet[column_letter]:
            try:
                value = cell.value
                if value is None:
                    continue
                length = len(str(value))
                if length > max_length:
                    max_length = length
            except Exception:
                continue
        worksheet.column_dimensions[column_letter].width = min(max_length + 4, 40)


def _build_operation_expression(
    operator: str,
    base_ids: Optional[List[int]],
    attr_ids: Optional[List[int]],
    row_idx: int,
    column_map: Dict[Tuple[HeaderKind, int], str],
) -> Optional[str]:
    references: List[str] = []
    for base_id in base_ids or []:
        column_letter = column_map.get(("base", int(base_id)))
        if column_letter:
            references.append(f"{column_letter}{row_idx}")
    for attr_id in attr_ids or []:
        column_letter = column_map.get(("atribute", int(attr_id)))
        if column_letter:
            references.append(f"{column_letter}{row_idx}")

    if not references:
        return None

    if operator == "multiplicacion":
        return "*".join(references)
    if operator == "division":
        return "/".join(references)
    if operator == "suma":
        return "+".join(references)
    if operator == "resta":
        return "-".join(references)
    return None


def _combine_formula(existing: str, new_expr: str, operator: str) -> str:
    if operator == "multiplicacion":
        return f"({existing})*({new_expr})"
    if operator == "division":
        return f"({existing})/({new_expr})"
    if operator == "suma":
        return f"({existing})+({new_expr})"
    if operator == "resta":
        return f"({existing})-({new_expr})"
    return existing


def _build_formula_for_header(
    row_idx: int,
    header: HeaderSpec,
    column_map: Dict[Tuple[HeaderKind, int], str],
) -> Optional[str]:
    calculo = header.calculo or {}
    if not calculo.get("activo"):
        return None
    operaciones = calculo.get("operaciones") or []
    if not operaciones:
        return None

    ops_iterable = operaciones if calculo.get("isMultiple") else operaciones[:1]

    formula_expr: Optional[str] = None

    for operacion in ops_iterable:
        operator = (operacion.get("tipo") or "multiplicacion").lower()
        expr = _build_operation_expression(
            operator,
            operacion.get("headers_base"),
            operacion.get("headers_atributes"),
            row_idx,
            column_map,
        )
        if not expr:
            continue
        if formula_expr is None:
            formula_expr = expr
        else:
            formula_expr = _combine_formula(formula_expr, expr, operator)

    if not formula_expr:
        return None

    return f"={formula_expr}"


def build_excel_for_tipo_material(tipo: TipoMaterial, materiales: List[Material]) -> bytes:
    headers = _build_ordered_headers(tipo)
    if not headers:
        headers = [
            HeaderSpec(
                kind="base",
                header_id=1,
                titulo="Detalle",
                calculo={},
                is_cantidad=False,
                order=1,
                raw={"id_header_base": 1, "titulo": "Detalle"},
            )
        ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = (tipo.titulo or "Materiales")[:31] or "Materiales"

    column_count = len(headers)
    title_cell = worksheet.cell(row=1, column=1, value=tipo.titulo)
    _apply_title_style(title_cell)
    if column_count > 1:
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)

    column_map = {
        (header.kind, header.header_id): get_column_letter(idx)
        for idx, header in enumerate(headers, start=1)
    }
    headers_lookup = {(header.kind, header.header_id): header for header in headers}

    # Crear headers en fila 2
    for idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=2, column=idx, value=header.titulo or "")
        _apply_header_style(cell)

    # Identificar columnas con cálculo activo
    columns_with_formulas: Dict[int, HeaderSpec] = {}
    for col_idx, header in enumerate(headers, start=1):
        calculo = header.calculo or {}
        if calculo.get("activo"):
            columns_with_formulas[col_idx] = header

    totals_map: Dict[Tuple[HeaderKind, int], float] = {}

    # Llenar datos de materiales existentes (filas 3 en adelante)
    for row_idx, material in enumerate(materiales, start=3):
        values_map = _build_row_values(tipo, material, headers)
        for col_idx, header in enumerate(headers, start=1):
            value = values_map.get((header.kind, header.header_id))
            cell = worksheet.cell(row=row_idx, column=col_idx)
            
            # Si esta columna tiene cálculo, NO poner fórmula aquí
            # (las fórmulas se pondrán después en todas las filas)
            if col_idx in columns_with_formulas:
                # Dejar celda vacía por ahora, se llenará con fórmulas después
                _apply_body_style(cell)
                numerical_for_total = _safe_to_float(value)
                totals_map[(header.kind, header.header_id)] = (
                    totals_map.get((header.kind, header.header_id), 0.0) + numerical_for_total
                )
                continue

            # Para columnas sin cálculo, llenar con valores
            if header.numeric_hint:
                numeric_value = _safe_to_float(value)
                cell.value = numeric_value
                totals_map[(header.kind, header.header_id)] = (
                    totals_map.get((header.kind, header.header_id), 0.0) + numeric_value
                )
            else:
                cell.value = value if value is not None else ""
            _apply_body_style(cell)

    # Ahora aplicar fórmulas a TODAS las filas (desde fila 3 hasta MAX_FORMULA_ROWS) 
    # para las columnas que tienen cálculo
    for col_idx, header in columns_with_formulas.items():
        for row_idx in range(3, MAX_FORMULA_ROWS + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            formula = _build_formula_for_header(row_idx, header, column_map)
            if formula:
                cell.value = formula
                _apply_body_style(cell)

    _auto_adjust_columns(worksheet, column_count)

    # Tabla de totales (a la derecha)
    totals_start_column = column_count + 3
    totals_title_cell = worksheet.cell(
        row=1,
        column=totals_start_column,
        value="Tabla de totales",
    )
    _apply_title_style(totals_title_cell)
    worksheet.merge_cells(
        start_row=1,
        start_column=totals_start_column,
        end_row=1,
        end_column=totals_start_column + 1,
    )

    # Construir mapa inverso: header_id -> columna
    header_to_column: Dict[Tuple[HeaderKind, int], str] = {}
    for idx, header in enumerate(headers, start=1):
        header_to_column[(header.kind, header.header_id)] = get_column_letter(idx)

    total_cantidad_struct = ensure_total_cantidad_struct(tipo.total_cantidad)
    cantidad_entries = total_cantidad_struct.get("cantidades") or []

    headers_lookup = {(spec.kind, spec.header_id): spec for spec in headers}

    # Lista de filas para la tabla de totales con etiqueta, fórmula/valor y referencia de celda
    totals_rows: List[Tuple[str, Any, Optional[str]]] = []
    total_cantidades_cells: List[str] = []  # Referencias de celdas para sumar Total Cantidades

    # Fila 2: Costo Unitario (header base id=4)
    costo_unitario_column = header_to_column.get(("base", 4))
    if costo_unitario_column:
        formula_costo_unitario = f"=SUM({costo_unitario_column}3:{costo_unitario_column}{MAX_FORMULA_ROWS})"
        totals_rows.append(("Costo Unitario", formula_costo_unitario, None))
    else:
        totals_rows.append(("Costo Unitario", 0.0, None))

    # Fila 3: Costo Total (header base id=5)
    costo_total_column = header_to_column.get(("base", 5))
    costo_total_row_ref = None
    if costo_total_column:
        formula_costo_total = f"=SUM({costo_total_column}3:{costo_total_column}{MAX_FORMULA_ROWS})"
        current_row = 2 + len(totals_rows)
        costo_total_row_ref = f"{get_column_letter(totals_start_column + 1)}{current_row}"
        totals_rows.append(("Costo Total", formula_costo_total, costo_total_row_ref))
    else:
        totals_rows.append(("Costo Total", 0.0, None))

    # Filas para cada entrada de cantidad
    for entry in cantidad_entries:
        header_type = _normalize_header_type(entry.get("typeOfHeader"))
        try:
            header_id = int(entry.get("idHeader", 0))
        except (TypeError, ValueError):
            continue
        if header_type == "base" and header_id in {4, 5}:
            continue
        
        header_spec = headers_lookup.get((header_type, header_id))
        if not header_spec and header_type == "base" and header_id == 2:
            header_label = "Cantidad"
        elif header_spec:
            header_label = header_spec.titulo or ("Header" if header_type == "atribute" else f"Base {header_id}")
        else:
            header_label = f"Header {header_id}"
        
        cantidad_column = header_to_column.get((header_type, header_id))
        if cantidad_column:
            formula_cantidad = f"=SUM({cantidad_column}3:{cantidad_column}{MAX_FORMULA_ROWS})"
            current_row = 2 + len(totals_rows)
            cell_ref = f"{get_column_letter(totals_start_column + 1)}{current_row}"
            totals_rows.append((f"Total {header_label}", formula_cantidad, cell_ref))
            total_cantidades_cells.append(cell_ref)
        else:
            totals_rows.append((f"Total {header_label}", 0.0, None))

    # Total costo cantidades (solo si hay más de una cantidad)
    total_cantidades_row_ref = None
    if len(cantidad_entries) > 1:
        if total_cantidades_cells:
            formula_total_cantidades = f"={'+'.join(total_cantidades_cells)}"
            current_row = 2 + len(totals_rows)
            total_cantidades_row_ref = f"{get_column_letter(totals_start_column + 1)}{current_row}"
            totals_rows.append(("Total costo cantidades", formula_total_cantidades, total_cantidades_row_ref))
        else:
            totals_rows.append(("Total costo cantidades", 0.0, None))

    # Separar el valor del dólar en su propia celda
    valor_dolar_column = totals_start_column + 2
    valor_dolar_label_cell = worksheet.cell(
        row=1,
        column=valor_dolar_column,
        value="Valor del dólar:",
    )
    _apply_body_style(valor_dolar_label_cell)
    
    valor_dolar_value_cell = worksheet.cell(
        row=1,
        column=valor_dolar_column + 1,
        value=tipo.valor_dolar,
    )
    _apply_body_style(valor_dolar_value_cell)
    valor_dolar_cell_ref = f"{get_column_letter(valor_dolar_column + 1)}1"

    # Total USD
    if costo_total_row_ref and valor_dolar_cell_ref:
        formula_total_usd = f"={costo_total_row_ref}*{valor_dolar_cell_ref}"
        totals_rows.append(("Total USD", formula_total_usd, None))
    else:
        totals_rows.append(("Total USD", tipo.total_USD, None))

    # Escribir las filas de totales
    for offset, (label, value, cell_ref) in enumerate(totals_rows, start=2):
        label_cell = worksheet.cell(row=offset, column=totals_start_column, value=label)
        value_cell = worksheet.cell(row=offset, column=totals_start_column + 1, value=value)
        _apply_body_style(label_cell)
        _apply_body_style(value_cell)
        label_cell.fill = PatternFill(fill_type="solid", fgColor=TOTAL_LABEL_FILL_COLOR)
        value_cell.fill = PatternFill(fill_type="solid", fgColor=TOTAL_VALUE_FILL_COLOR)

    _auto_adjust_columns(worksheet, valor_dolar_column + 1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


__all__ = ["build_excel_for_tipo_material"]
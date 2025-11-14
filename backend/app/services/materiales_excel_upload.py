from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple
from io import BytesIO

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import Material, TipoMaterial, ensure_total_cantidad_struct


def _safe_to_float(value: Any) -> float:
    """Convierte un valor de Excel a float de forma segura"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Remover símbolos de moneda y espacios
        cleaned = value.strip().replace('$', '').replace(',', '.')
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0


def _increment_quantity_total(
    quantity_totals: Dict[tuple[str, int], float],
    tipo_header: str,
    header_id: int,
    value: Any,
) -> None:
    numeric = _safe_to_float(value)
    if numeric == 0.0:
        return
    key = (tipo_header, header_id)
    quantity_totals[key] = quantity_totals.get(key, 0.0) + numeric


def _normalize_header_title(title: str) -> str:
    """Normaliza el título de un header para comparación"""
    return title.strip().lower()


def _find_header_column(
    tipo: TipoMaterial,
    normalized_title: str
) -> Optional[tuple[str, int]]:
    """
    Encuentra la columna correspondiente a un header.
    Retorna tupla (tipo, id) donde tipo es 'base' o 'atribute'
    """
    # Buscar en headers base
    for header in tipo.headers_base or []:
        if not header.get('active', True):
            continue
        header_title = _normalize_header_title(header.get('titulo', ''))
        if header_title == normalized_title:
            return ('base', header['id_header_base'])
    
    # Buscar en headers atributos
    for header in tipo.headers_atributes or []:
        header_title = _normalize_header_title(header.get('titulo', ''))
        if header_title == normalized_title:
            return ('atribute', header['id_header_atribute'])
    
    return None


def _extract_valor_dolar_from_totals(worksheet, totals_start_column: int) -> Optional[float]:
    """Extrae el valor del dólar de la tabla de totales"""
    # El valor del dólar está en la fila 1, columnas totals_start_column+2 y +3
    # Columna +2 tiene la etiqueta "Valor del dólar:"
    # Columna +3 tiene el valor numérico
    try:
        valor_dolar_cell = worksheet.cell(row=1, column=totals_start_column + 3)
        valor = _safe_to_float(valor_dolar_cell.value)
        return valor if valor > 0 else None
    except Exception:
        return None


def _read_materials_from_excel(
    worksheet,
    tipo: TipoMaterial,
    column_count: int
) -> tuple[List[Dict[str, Any]], Dict[tuple[str, int], float]]:
    """
    Lee los materiales del Excel y los convierte al formato esperado por el backend
    """
    base_map = {h['id_header_base']: h for h in tipo.headers_base or []}
    attr_map = {h['id_header_atribute']: h for h in tipo.headers_atributes or []}

    # Construir mapa de columnas
    header_row = 2
    column_map: Dict[int, Dict[str, Any]] = {}
    
    for col_idx in range(1, column_count + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        title = str(cell.value or '').strip()
        if not title:
            continue
        
        normalized_title = _normalize_header_title(title)
        header_info = _find_header_column(tipo, normalized_title)
        
        if header_info:
            tipo_header, header_id = header_info
            if tipo_header == 'base':
                header_meta = base_map.get(header_id) or {}
                is_cantidad = header_id == 2
            else:
                header_meta = attr_map.get(header_id) or {}
                is_cantidad = bool(header_meta.get('isCantidad'))

            calculo_activo = bool((header_meta.get('calculo') or {}).get('activo'))

            column_map[col_idx] = {
                "type": tipo_header,
                "id": header_id,
                "title": title,
                "isCantidad": is_cantidad,
                "calculoActivo": calculo_activo,
            }
    
    # Leer materiales desde la fila 3 en adelante
    materials: List[Dict[str, Any]] = []
    quantity_totals: Dict[tuple[str, int], float] = {}
    
    # Determinar hasta qué fila leer (buscar la primera fila completamente vacía)
    max_row = worksheet.max_row
    last_data_row = 3
    
    for row_idx in range(3, max_row + 1):
        has_data = False
        for col_idx in range(1, column_count + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None and str(cell_value).strip():
                has_data = True
                break
        
        if has_data:
            last_data_row = row_idx
        else:
            # Si encontramos 5 filas vacías consecutivas, paramos
            if row_idx - last_data_row >= 5:
                break
    
    # Leer datos de materiales
    for row_idx in range(3, last_data_row + 1):
        material_data: Dict[str, Any] = {
            'detalle': '',
            'unidad': None,
            'cantidad': None,
            'costo_unitario': 0.0,
            'costo_total': 0.0,
            'atributos': []
        }

        has_content = False

        for col_idx, col_info in column_map.items():
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = cell.value

            if value is not None and str(value).strip():
                has_content = True

            tipo_header = col_info["type"]
            header_id = col_info["id"]

            if tipo_header == 'base':
                if header_id == 1:
                    material_data['detalle'] = str(value or '').strip()
                elif header_id == 2:
                    material_data['cantidad'] = str(value or '').strip()
                    if col_info.get("isCantidad"):
                        _increment_quantity_total(quantity_totals, 'base', header_id, value)
                elif header_id == 3:
                    material_data['unidad'] = str(value or '').strip()
                elif header_id == 4:
                    material_data['costo_unitario'] = _safe_to_float(value)
                elif header_id == 5:
                    material_data['costo_total'] = _safe_to_float(value)
                else:
                    header = base_map.get(header_id)
                    if header and col_info.get("isCantidad"):
                        _increment_quantity_total(quantity_totals, 'base', header_id, value)
            elif tipo_header == 'atribute':
                header = attr_map.get(header_id)
                if not header:
                    continue

                material_data['atributos'].append({
                    'id_header_atribute': header_id,
                    'value': str(value or '').strip()
                })

                if col_info.get("isCantidad"):
                    _increment_quantity_total(quantity_totals, 'atribute', header_id, value)

        if has_content and material_data['detalle']:
            materials.append(material_data)

    return materials, quantity_totals


def _extract_totals_from_excel(
    worksheet,
    tipo: TipoMaterial,
    totals_start_column: int
) -> Dict[str, float]:
    """
    Extrae los totales de la tabla de totales del Excel
    """
    totals: Dict[str, float] = {
        'total_costo_unitario': 0.0,
        'total_costo_total': 0.0,
        'total_USD': 0.0,
        'total_cantidades': 0.0,
    }
    
    # Mapa de etiquetas a keys
    label_map = {
        'costo unitario': 'total_costo_unitario',
        'costo total': 'total_costo_total',
        'total usd': 'total_USD',
        'total costo cantidades': 'total_cantidades',
    }
    
    # Leer desde la fila 2 en adelante
    for row_idx in range(2, 20):  # Leer hasta 20 filas como máximo
        try:
            label_cell = worksheet.cell(row=row_idx, column=totals_start_column)
            value_cell = worksheet.cell(row=row_idx, column=totals_start_column + 1)
            
            label = str(label_cell.value or '').strip().lower()
            if not label:
                continue
            
            if label in label_map:
                key = label_map[label]
                totals[key] = _safe_to_float(value_cell.value)
        except Exception:
            continue
    
    return totals


def _apply_global_valor_dolar(db: Session, nuevo_valor: float) -> None:
    """Actualiza el valor del dólar en todos los tipos de material"""
    tipos = db.scalars(select(TipoMaterial)).all()
    for tipo in tipos:
        tipo.valor_dolar = float(nuevo_valor)
        tipo.total_USD = float(tipo.total_costo_total or 0.0) * float(nuevo_valor)
        db.add(tipo)


async def process_excel_upload(
    file: UploadFile,
    id_tipo_material: int,
    db: Session
) -> Dict[str, Any]:
    """
    Procesa un archivo Excel subido y actualiza la base de datos:
    1. Borra todos los materiales existentes del tipo
    2. Carga los nuevos materiales del Excel
    3. Actualiza los totales del tipo de material
    4. Si el valor del dólar cambió, actualiza todos los tipos de material
    """
    
    # Verificar que el tipo de material existe
    tipo = db.get(TipoMaterial, id_tipo_material)
    if not tipo:
        raise HTTPException(status_code=404, detail="Tipo de material no encontrado")
    
    # Leer el archivo Excel
    try:
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents), data_only=True)
        worksheet = workbook.active
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error al leer el archivo Excel: {str(e)}"
        )
    
    # Determinar el número de columnas de la tabla de materiales
    # La tabla comienza en la fila 1 (título) y fila 2 (headers)
    column_count = 0
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=2, column=col_idx)
        if cell.value:
            column_count = col_idx
        else:
            break
    
    if column_count == 0:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron headers en el Excel"
        )
    
    # La tabla de totales comienza 3 columnas después
    totals_start_column = column_count + 3
    
    # Extraer valor del dólar
    nuevo_valor_dolar = _extract_valor_dolar_from_totals(worksheet, totals_start_column)
    valor_dolar_cambio = False
    
    if nuevo_valor_dolar and abs(nuevo_valor_dolar - tipo.valor_dolar) > 0.01:
        valor_dolar_cambio = True
        _apply_global_valor_dolar(db, nuevo_valor_dolar)
        db.flush()
    
    # Extraer totales
    totals = _extract_totals_from_excel(worksheet, tipo, totals_start_column)
    
    # Leer materiales del Excel
    materials_data, quantity_totals = _read_materials_from_excel(worksheet, tipo, column_count)
    
    if not materials_data:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron materiales en el Excel"
        )
    
    # Borrar todos los materiales existentes del tipo
    stmt = select(Material).where(Material.id_tipo_material == id_tipo_material)
    existing_materials = db.scalars(stmt).all()
    
    for material in existing_materials:
        db.delete(material)
    
    db.flush()
    
    # Crear los nuevos materiales
    created_materials: List[Material] = []
    
    for material_data in materials_data:
        new_material = Material(
            id_tipo_material=id_tipo_material,
            detalle=material_data['detalle'],
            unidad=material_data.get('unidad'),
            cantidad=material_data.get('cantidad'),
            costo_unitario=material_data.get('costo_unitario', 0.0),
            costo_total=material_data.get('costo_total', 0.0),
            atributos=material_data.get('atributos', [])
        )
        db.add(new_material)
        created_materials.append(new_material)
    
    # Actualizar totales del tipo de material
    tipo.total_costo_unitario = totals.get('total_costo_unitario', 0.0)
    tipo.total_costo_total = totals.get('total_costo_total', 0.0)
    tipo.total_USD = totals.get('total_USD', 0.0)
    
    # Construir estructura de cantidades
    cantidades_entries: List[Dict[str, Any]] = []

    for header in tipo.headers_base or []:
        if header.get('id_header_base') == 2 and header.get('active', True):
            total = quantity_totals.get(('base', 2), 0.0)
            cantidades_entries.append({
                "typeOfHeader": "base",
                "idHeader": 2,
                "total": total,
            })
            break

    for header in tipo.headers_atributes or []:
        if header.get('isCantidad'):
            header_id = header['id_header_atribute']
            total = quantity_totals.get(('atribute', header_id), 0.0)
            cantidades_entries.append({
                "typeOfHeader": "atribute",
                "idHeader": header_id,
                "total": total,
            })

    total_cantidades_sum = sum(entry.get("total", 0.0) for entry in cantidades_entries)
    tipo.total_cantidad = ensure_total_cantidad_struct(
        {
            "total_cantidades": total_cantidades_sum,
            "cantidades": cantidades_entries,
        }
    )
    
    db.add(tipo)
    
    # Commit de todos los cambios
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error al guardar los datos: {str(e)}"
        )
    
    # Refrescar el tipo para obtener los datos actualizados
    db.refresh(tipo)
    
    return {
        'success': True,
        'materiales_creados': len(created_materials),
        'valor_dolar_actualizado': valor_dolar_cambio,
        'nuevo_valor_dolar': nuevo_valor_dolar if valor_dolar_cambio else None,
        'totales_actualizados': {
            'total_costo_unitario': tipo.total_costo_unitario,
            'total_costo_total': tipo.total_costo_total,
            'total_USD': tipo.total_USD,
        }
    }


__all__ = ['process_excel_upload']